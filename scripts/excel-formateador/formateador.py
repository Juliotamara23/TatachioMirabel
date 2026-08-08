#!/usr/bin/env python3
"""
Formateador de Censos — Excel del Ministerio del Interior.

Portado al monorepo TatachioMirabel desde AnalisisCensal
(https://github.com/Juliotamara23/AnalisisCensal). Mantiene la lógica
probada: copia la plantilla oficial del Ministerio preservando logos y
estructura, y solo inyecta las filas de datos con openpyxl celda a celda.

Uso:
    python scripts/excel-formateador/formateador.py \
        --origen <datos.xlsx> \
        --plantilla <plantilla-ministerio.xlsx> \
        --destino <salida.xlsx>

El archivo de origen puede venir de un export del backend (DB -> Excel)
o de un cuestionario externo. La plantilla es el formato oficial del
Ministerio del Interior (NO se sube al repo por ser propiedad ministerial).

Requiere: pip install -r scripts/excel-formateador/requirements.txt
"""

import argparse
import os
import shutil
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Tipos de datos y mapeos oficiales del Ministerio del Interior
TIPOS_ESPERADOS = {
    "VIGENCIA": {"tipo": "año", "mapeo": None},
    "RESGUARDO INDIGENA": {"tipo": "texto", "mapeo": None},
    "COMUNIDAD INDIGENA": {"tipo": "texto", "mapeo": None},
    "FAMILIA": {"tipo": "numero", "mapeo": None},
    "TIPO IDENTIFICACION": {"tipo": "codigo", "mapeo": {
        "Cédula de Ciudadanía": "CC", "CC": "CC",
        "Tarjeta de Identidad": "TI", "TI": "TI",
        "Registro Civil de Nacimiento": "RC", "RC": "RC",
        "NUIP": "NUIP", "Numero Único": "NUIP"
    }},
    "NUMERO DOCUMENTO": {"tipo": "texto_limpio", "mapeo": None},
    "NOMBRES": {"tipo": "mayusculas", "mapeo": None},
    "APELLIDOS": {"tipo": "mayusculas", "mapeo": None},
    "FECHA NACIMIENTO": {"tipo": "fecha", "mapeo": "%d/%m/%Y"},
    "PARENTESCO": {"tipo": "codigo", "mapeo": {
        "Padre": "PA", "PA": "PA", "Madre": "MA", "MA": "MA",
        "Cónyuge": "CO", "CO": "CO", "Esposo": "ES", "Esposa": "ES", "ES": "ES",
        "Hermano(a)": "HE", "HE": "HE",
        "Cabeza de Familia": "CF", "Jefe": "CF", "CF": "CF",
        "Hijo": "HI", "HI": "HI", "Hijo(a)": "HI",
        "Yerno": "YR", "YR": "YR", "Nuera": "NU", "NU": "NU",
        "Suegro": "SU", "SU": "SU",
        "Sobrino": "SO", "SO": "SO",
        "Cuñado": "CU", "CU": "CU",
        "Tío": "TI", "TI": "TI",
        "Abuelo": "AB", "AB": "AB",
        "Nieto": "NI", "NI": "NI", "Nieta": "NI"
    }},
    "SEXO": {"tipo": "codigo", "mapeo": {
        "Masculino": "M", "M": "M",
        "Femenino": "F", "F": "F"
    }},
    "ESTADO CIVIL": {"tipo": "codigo", "mapeo": {
        "Soltero": "S", "S": "S", "Soltero(a)": "S",
        "Casado": "C", "C": "C", "Casado(a)": "C", "Unión libre": "C",
        "Divorciado": "S", "Viudo": "S"
    }},
    "PROFESION": {"tipo": "mayusculas", "mapeo": None},
    "ESCOLARIDAD": {"tipo": "codigo", "mapeo": {
        "PR": "PR", "SE": "SE", "UN": "UN", "NI": "NI",
        "Primaria": "PR", "Secundaria": "SE", "Universitaria": "UN", "Ninguno": "NI"
    }},
    "INTEGRANTES": {"tipo": "numero", "mapeo": None},
    "DIRECCION": {"tipo": "mayusculas", "mapeo": None},
    "TELEFONO": {"tipo": "telefono", "mapeo": None},
    "USUARIO": {"tipo": "mayusculas", "mapeo": None},
}


def encontrar_fila_encabezados(ruta_archivo):
    """Escanea las primeras filas para encontrar los encabezados del Ministerio."""
    keywords = ["VIGENCIA", "RESGUARDO", "COMUNIDAD", "FAMILIA", "IDENTIFICACION", "DOCUMENTO"]
    for i in range(15):
        try:
            df_temp = pd.read_excel(ruta_archivo, header=i, nrows=1)
            cols = [str(c).upper() for c in df_temp.columns]
            coincidencias = sum(1 for kw in keywords if any(kw in c for c in cols))
            if coincidencias >= 3:
                return i + 1
        except Exception:
            continue
    return -1


def limpiar_valor(valor, tipo_info):
    """Limpia y formatea un valor según el tipo esperado."""
    if valor is None:
        return ""

    valor_str = str(valor).strip()
    if valor_str.upper() in ["NAN", "NONE", ""]:
        return ""

    tipo = tipo_info["tipo"]
    mapeo = tipo_info.get("mapeo")

    if tipo == "año":
        try:
            return int(float(valor_str))
        except (ValueError, TypeError):
            return datetime.now().year

    if tipo == "mayusculas":
        return valor_str.upper()

    if tipo == "texto_limpio":
        return valor_str.replace(".", "").replace(" ", "").replace("-", "")

    if tipo == "telefono":
        telefono_limpio = valor_str.strip()
        if telefono_limpio.endswith(".0"):
            telefono_limpio = telefono_limpio[:-2]
        return "".join(c for c in telefono_limpio if c.isdigit())

    if tipo == "codigo":
        if mapeo:
            val_upper = valor_str.upper()
            for k, v in mapeo.items():
                if k.upper() == val_upper or k.upper() in val_upper.upper():
                    return v
            if valor_str in mapeo.values():
                return valor_str
        return valor_str

    if tipo == "fecha":
        try:
            dt = pd.to_datetime(valor, dayfirst=True)
            return dt.strftime("%d/%m/%Y")
        except (ValueError, TypeError):
            return valor_str

    if tipo == "numero":
        try:
            return int(float(valor_str))
        except (ValueError, TypeError):
            return valor_str

    return valor_str


def obtener_headers(ruta_archivo, fila_header):
    """Obtiene los encabezados de un archivo excel."""
    df = pd.read_excel(ruta_archivo, header=fila_header - 1)
    df.columns = df.columns.astype(str).str.strip().str.upper()
    return df


def validar_archivo(ruta_origen, ruta_referencia):
    """Valida que el archivo origen sea compatible con el formato de referencia."""
    print("=" * 60)
    print("FASE 1: ANALISIS DE COMPATIBILIDAD")
    print("=" * 60)

    if not os.path.exists(ruta_origen):
        return False, f"Error: El archivo origen '{ruta_origen}' no fue encontrado.", None, None

    if not os.path.exists(ruta_referencia):
        return False, f"Error: El archivo de referencia '{ruta_referencia}' no fue encontrado.", None, None

    fila_origen = encontrar_fila_encabezados(ruta_origen)
    fila_ref = encontrar_fila_encabezados(ruta_referencia)

    if fila_origen == -1:
        return False, "Error: No se detectaron encabezados validos en el archivo origen.", None, None

    if fila_ref == -1:
        return False, "Error: No se detectaron encabezados validos en el archivo de referencia.", None, None

    print(f"-> Encabezados detectados en origen: Fila {fila_origen}")
    print(f"-> Encabezados detectados en referencia: Fila {fila_ref}")

    df_origen = obtener_headers(ruta_origen, fila_origen)
    df_ref = obtener_headers(ruta_referencia, fila_ref)

    headers_ref = df_ref.columns.tolist()

    print(f"\n-> Headers del archivo de referencia ({len(headers_ref)}):")
    for h in headers_ref:
        print(f"   - {h}")

    print("\n-> Validando compatibilidad de columnas...")
    faltantes = []
    for header_ref in headers_ref:
        encontrado = False
        header_ref_upper = header_ref.upper()

        for col_origen in df_origen.columns:
            col_origen_upper = col_origen.upper()
            if header_ref_upper == col_origen_upper or \
               header_ref_upper in col_origen_upper or \
               col_origen_upper in header_ref_upper:
                encontrado = True
                break

        if not encontrado:
            faltantes.append(header_ref)
            print(f"   [X] FALTA: '{header_ref}'")
        else:
            print(f"   [OK] '{header_ref}'")

    if faltantes:
        return False, f"Error: El archivo origen no es compatible. Faltan columnas: {', '.join(faltantes)}", None, None

    df_datos = pd.read_excel(ruta_origen, header=fila_origen - 1)
    df_datos.columns = df_datos.columns.astype(str).str.strip()

    df_datos = df_datos.dropna(how="all")
    df_datos = df_datos[df_datos.apply(lambda row: row.notna().any(), axis=1)]

    print(f"\n-> Registros detectados: {len(df_datos)}")

    mapeo_columnas = {}
    for header_ref in headers_ref:
        header_ref_upper = header_ref.upper()
        for col_origen in df_datos.columns:
            col_origen_upper = col_origen.upper()
            if header_ref_upper == col_origen_upper or \
               header_ref_upper in col_origen_upper or \
               col_origen_upper in header_ref_upper:
                mapeo_columnas[header_ref] = col_origen
                break

    return True, "OK", df_datos, mapeo_columnas


def ejecutar_formateo(ruta_origen, ruta_destino, ruta_referencia):
    """Ejecuta el proceso completo de formateo."""
    es_compatible, mensaje, df_datos, mapeo_col = validar_archivo(ruta_origen, ruta_referencia)

    if not es_compatible:
        print(f"\n[MENSAJE] {mensaje}")
        return False

    print("\n" + "=" * 60)
    print("FASE 2: TRANSFORMACION DE DATOS")
    print("=" * 60)

    num_filas = len(df_datos)
    df_formateado = pd.DataFrame(index=range(num_filas))

    for col_ref, tipo_info in TIPOS_ESPERADOS.items():
        col_origen = mapeo_col.get(col_ref)

        if col_origen and col_origen in df_datos.columns:
            print(f"-> Transformando '{col_ref}' desde '{col_origen}'...")
            valores = df_datos[col_origen].apply(lambda x: limpiar_valor(x, tipo_info))
            df_formateado[col_ref] = valores.values
        else:
            print(f"-> Valor por defecto para '{col_ref}'...")
            if col_ref == "VIGENCIA":
                df_formateado[col_ref] = datetime.now().year
            elif col_ref == "RESGUARDO INDIGENA":
                df_formateado[col_ref] = "0"
            elif col_ref == "COMUNIDAD INDIGENA":
                df_formateado[col_ref] = "TATACHIO MIRABEL"
            elif col_ref == "USUARIO":
                df_formateado[col_ref] = "SISTEMA"
            elif col_ref == "ESCOLARIDAD":
                df_formateado[col_ref] = "NI"
            else:
                df_formateado[col_ref] = ""

    orden = list(TIPOS_ESPERADOS.keys())
    df_formateado = df_formateado.reindex(columns=orden)

    print(f"\n-> Datos transformados: {len(df_formateado)} filas")

    print("\n" + "=" * 60)
    print("FASE 3: INYECCION EN PLANTILLA")
    print("=" * 60)

    shutil.copy2(ruta_referencia, ruta_destino)
    print(f"-> Creada copia de referencia: {os.path.basename(ruta_destino)}")

    fila_ref = encontrar_fila_encabezados(ruta_referencia)
    fila_inicio_datos = fila_ref + 1

    print(f"-> Encabezados en fila {fila_ref}, datos comienzan en fila {fila_inicio_datos + 1}")

    wb = load_workbook(ruta_destino)
    ws = wb.active

    filas_totales = ws.max_row if ws.max_row else 1000
    for row in range(fila_inicio_datos, filas_totales + 1):
        for col in range(1, 19):
            ws.cell(row=row, column=col, value=None)

    for r_idx, row_data in enumerate(df_formateado.values, start=fila_inicio_datos):
        for c_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="center")

    wb.save(ruta_destino)
    wb.close()

    print(f"\n" + "=" * 60)
    print("[OK] PROCESO COMPLETADO CON EXITO")
    print("=" * 60)
    print(f"-> Archivo de salida: {ruta_destino}")
    print(f"-> Total de registros: {len(df_formateado)}")

    return True


def main():
    parser = argparse.ArgumentParser(description="Formateador de censos para el Ministerio del Interior")
    parser.add_argument("--origen", required=True, help="Archivo Excel con datos de origen")
    parser.add_argument("--plantilla", required=True, help="Plantilla oficial del Ministerio (xlsx)")
    parser.add_argument("--destino", required=True, help="Archivo de salida (xlsx)")
    args = parser.parse_args()

    print("=" * 60)
    print(f"   SCRIPT DE FORMATEO DE CENSOS - {datetime.now().year}")
    print("=" * 60)

    ok = ejecutar_formateo(args.origen, args.destino, args.plantilla)
    if not ok:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
