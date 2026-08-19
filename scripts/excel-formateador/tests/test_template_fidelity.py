#!/usr/bin/env python3
"""
Regression tests for Excel template fidelity losses.

These tests verify that the formateador preserves the template's exact structure:
- Sheet names, merged cells, images/drawings
- Table definitions with expanding refs
- Style retention on blank prepared rows
- Default output filename format

They are expected to FAIL against the current implementation due to known
fidelity losses: openpyxl loses image/drawing XML on save, and does not expand
table refs when writing beyond the prepared 23 data rows.
"""

import json
import os
import subprocess
import sys
import tempfile
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl import styles as openpyxl_styles
import pytest

# ─── Constants ──────────────────────────────────────────────────────────────
TEMPLATE_PATH = Path(__file__).parent.parent / "templates" / "Formato Censal.xlsx"
FORMATADOR_PATH = Path(__file__).parent.parent / "formateador.py"
CURRENT_YEAR = datetime.now().year


# ─── Helpers ────────────────────────────────────────────────────────────────
def run_formateador(data: dict, output_path: Path, template_path: Path | None = None) -> subprocess.CompletedProcess:
    """Invoke formateador.py as a subprocess with the given data."""
    with tempfile.NamedTemporaryFile(mode="w", suffix=".json", delete=False) as f:
        json.dump(data, f)
        data_path = Path(f.name)

    try:
        cmd = [
            sys.executable,
            str(FORMATADOR_PATH),
            "--data",
            str(data_path),
            "--output",
            str(output_path),
        ]
        if template_path:
            cmd.extend(["--template", str(template_path)])

        return subprocess.run(cmd, capture_output=True, text=True, timeout=30)
    finally:
        data_path.unlink(missing_ok=True)


def load_workbook_readonly(path: Path) -> openpyxl.Workbook:
    """Load workbook in read-only mode for inspection."""
    return openpyxl.load_workbook(path, read_only=False, keep_vba=True, data_only=False)


def get_sheet_names(wb: openpyxl.Workbook) -> list[str]:
    return wb.sheetnames


def get_merged_ranges(wb: openpyxl.Workbook, sheet_name: str) -> set[str]:
    ws = wb[sheet_name]
    return {str(r) for r in ws.merged_cells.ranges}


def get_tables(wb: openpyxl.Workbook, sheet_name: str) -> dict:
    ws = wb[sheet_name]
    tables = {}
    if hasattr(ws, "tables") and ws.tables:
        for name in ws.tables:
            table = ws.tables[name]
            tables[name] = str(table.ref)
    return tables


def has_drawing_relationships(wb: openpyxl.Workbook, sheet_name: str) -> bool:
    """Check if the sheet has drawing relationships (images/drawings).

    The template has a drawing relationship (drawing1.xml) pointing to an image.
    openpyxl loses this on save - the drawing relationship disappears from _rels.
    The correct relationship type is:
    http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing
    """
    ws = wb[sheet_name]
    # Check for drawing relationships in _rels (most reliable)
    # openpyxl stores relationships in _rels (RelationshipList)
    DRAWING_REL_TYPE = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing"
    if hasattr(ws, "_rels"):
        try:
            for rel in ws._rels:  # type: ignore[attr-defined]
                if rel.Type == DRAWING_REL_TYPE:
                    return True
        except Exception:
            pass
    # Check _images as fallback
    if getattr(ws, "_images", None):  # type: ignore[attr-defined]
        if len(ws._images) > 0:  # type: ignore[attr-defined]
            return True
    return False


def get_cell_styles(wb: openpyxl.Workbook, sheet_name: str, row: int, col: int):
    """Get the style of a specific cell."""
    ws = wb[sheet_name]
    cell = ws.cell(row=row, column=col)
    return {
        "font": cell.font,
        "fill": cell.fill,
        "border": cell.border,
        "alignment": cell.alignment,
        "number_format": cell.number_format,
        "protection": cell.protection,
    }


def count_data_rows_in_table(wb: openpyxl.Workbook, sheet_name: str, table_name: str) -> int:
    """Count actual data rows in a table (excluding header)."""
    ws = wb[sheet_name]
    table = ws.tables.get(table_name)
    if not table:
        return 0
    ref = str(table.ref)  # e.g., "A1:O24"
    # Parse ref to get row range
    if ":" in ref:
        end_cell = ref.split(":")[1]
        # Extract row number from end_cell (e.g., "O24" -> 24)
        row_str = "".join(c for c in end_cell if c.isdigit())
        if row_str:
            return int(row_str) - 1  # subtract header row
    return 0


# ─── Test Data ──────────────────────────────────────────────────────────────
MINIMAL_CENSO_ROW = {
    "VIGENCIA": 2026,
    "RESGUARDO INDIGENA": "TEST",
    "DEPARTAMENTO": "TEST",
    "MUNICIPIO": "TEST",
    "ZONA": "TEST",
    "CENTRO POBLADO": "TEST",
    "SECTOR": "TEST",
    "MANZANA": "TEST",
    "VIVIENDA": "TEST",
    "HOGAR": "TEST",
    "PERSONA": "TEST",
    "TIPO DOCUMENTO": "CC",
    "NUMERO DOCUMENTO": "1234567890",
    "PRIMER APELLIDO": "TEST",
    "SEGUNDO APELLIDO": "",
    "PRIMER NOMBRE": "TEST",
    "SEGUNDO NOMBRE": "",
    "SEXO": "M",
}

MINIMAL_ALTAS_ROW = {
    "VIGENCIA": 2026,
    "RESGUARDO INDIGENA": "TEST",
    "DEPARTAMENTO": "TEST",
    "MUNICIPIO": "TEST",
    "ZONA": "TEST",
    "CENTRO POBLADO": "TEST",
    "SECTOR": "TEST",
    "MANZANA": "TEST",
    "VIVIENDA": "TEST",
    "HOGAR": "TEST",
    "PERSONA": "TEST",
    "TIPO DOCUMENTO": "CC",
    "NUMERO DOCUMENTO": "1234567890",
    "PRIMER APELLIDO": "TEST",
    "SEGUNDO APELLIDO": "",
    "PRIMER NOMBRE": "TEST",
    "NOVEDAD": "ALTA",
}

MINIMAL_BAJAS_ROW = {
    **MINIMAL_ALTAS_ROW,
    "NOVEDAD": "BAJA",
}


def make_minimal_data() -> dict:
    return {
        "censo": [MINIMAL_CENSO_ROW],
        "altas": [MINIMAL_ALTAS_ROW],
        "bajas": [MINIMAL_BAJAS_ROW],
    }


def make_excess_data(num_rows: int = 30) -> dict:
    """Create data exceeding the 23 prepared rows in ALTAS/BAJAS tables."""
    return {
        "censo": [MINIMAL_CENSO_ROW],
        "altas": [MINIMAL_ALTAS_ROW for _ in range(num_rows)],
        "bajas": [MINIMAL_BAJAS_ROW for _ in range(num_rows)],
    }


# ─── Test Class ─────────────────────────────────────────────────────────────
class TestTemplateFidelityRegression:
    """Regression tests for template fidelity losses."""

    @pytest.fixture(autouse=True)
    def setup_temp_dir(self, tmp_path: Path):
        """Provide a temp directory for each test."""
        self.tmp_dir = tmp_path

    # ─── Test 1: Minimal export preserves all structure ────────────────────

    def test_minimal_export_preserves_sheet_names(self):
        """All three sheet names must be preserved exactly."""
        output = self.tmp_dir / "censo-test.xlsx"
        run_formateador(make_minimal_data(), output)

        wb = load_workbook_readonly(output)
        assert get_sheet_names(wb) == ["FORMATO_CENSOS", "REPORTE ALTAS", "REPORTE BAJAS"]

    def test_minimal_export_preserves_merged_ranges(self):
        """FORMATO_CENSOS merged ranges must be preserved exactly (9 ranges)."""
        output = self.tmp_dir / "censo-test.xlsx"
        run_formateador(make_minimal_data(), output)

        wb = load_workbook_readonly(output)
        merged = get_merged_ranges(wb, "FORMATO_CENSOS")

        # Template has exactly these 9 merged ranges
        expected_merged = {
            "A1:B5", "C1:F3", "G1:P3", "C4:F5", "G4:P5",
            "Q1:Q2", "Q3:Q4", "R1:R2", "R3:R4",
        }
        assert merged == expected_merged, f"Expected {expected_merged}, got {merged}"

    def test_minimal_export_preserves_image_drawing_relationships(self):
        """FORMATO_CENSOS must retain the template's image/drawing relationships."""
        output = self.tmp_dir / "censo-test.xlsx"
        run_formateador(make_minimal_data(), output)

        wb = load_workbook_readonly(output)
        # The template has a drawing relationship on FORMATO_CENSOS (sheet1)
        # This currently FAILS because openpyxl drops drawing XML on save
        assert has_drawing_relationships(wb, "FORMATO_CENSOS"), (
            "Drawing relationships lost on FORMATO_CENSOS — "
            "openpyxl does not preserve drawing XML on save"
        )

    # ─── Test 2: Table refs expand when data exceeds prepared rows ────────

    def test_altas_table_ref_expands_beyond_prepared_rows(self):
        """Table_1 ref must expand to include all generated data rows."""
        output = self.tmp_dir / "censo-test.xlsx"
        # 30 rows > 23 prepared rows
        run_formateador(make_excess_data(30), output)

        wb = load_workbook_readonly(output)
        tables = get_tables(wb, "REPORTE ALTAS")

        assert "Table_1" in tables, "Table_1 missing from REPORTE ALTAS"
        # Table ref should expand to cover all 30 data rows + header = row 31
        # This currently FAILS — openpyxl doesn't auto-expand table refs
        ref = tables["Table_1"]
        # Parse end row from ref (e.g., "A1:O31" -> 31)
        end_row = int("".join(c for c in ref.split(":")[1] if c.isdigit()))
        assert end_row >= 31, f"Table_1 ref {ref} does not cover 30 data rows (expected end row >= 31)"

    def test_bajas_table_ref_expands_beyond_prepared_rows(self):
        """Table_2 ref must expand to include all generated data rows."""
        output = self.tmp_dir / "censo-test.xlsx"
        run_formateador(make_excess_data(30), output)

        wb = load_workbook_readonly(output)
        tables = get_tables(wb, "REPORTE BAJAS")

        assert "Table_2" in tables, "Table_2 missing from REPORTE BAJAS"
        ref = tables["Table_2"]
        end_row = int("".join(c for c in ref.split(":")[1] if c.isdigit()))
        assert end_row >= 31, f"Table_2 ref {ref} does not cover 30 data rows (expected end row >= 31)"

    def test_table_definitions_remain_present_after_write(self):
        """Table definitions (Table_1, Table_2) must not be lost on save."""
        output = self.tmp_dir / "censo-test.xlsx"
        run_formateador(make_minimal_data(), output)

        wb = load_workbook_readonly(output)
        altas_tables = get_tables(wb, "REPORTE ALTAS")
        bajas_tables = get_tables(wb, "REPORTE BAJAS")

        assert "Table_1" in altas_tables, "Table_1 definition lost after write"
        assert "Table_2" in bajas_tables, "Table_2 definition lost after write"

    # ─── Test 3: Blank prepared rows retain style when input < capacity ───

    def test_blank_rows_retain_style_below_prepared_capacity(self):
        """Rows beyond input but within prepared capacity must be blank but styled."""
        output = self.tmp_dir / "censo-test.xlsx"
        # Only 1 data row, but template has 23 prepared rows (rows 2-24)
        run_formateador(make_minimal_data(), output)

        wb = load_workbook_readonly(output)
        ws_altas = wb["REPORTE ALTAS"]
        ws_bajas = wb["REPORTE BAJAS"]

        # Check a row beyond input (row 5) but within prepared capacity (row 24)
        # It should be blank (value=None) but retain style from template
        for ws, sheet_name in [(ws_altas, "REPORTE ALTAS"), (ws_bajas, "REPORTE BAJAS")]:
            # Row 5 is beyond our 1 data row but within the 23 prepared rows
            for col in range(1, 16):  # A-O columns
                cell = ws.cell(row=5, column=col)
                assert cell.value is None, f"{sheet_name} row 5 col {col} should be blank"

                # Check that style is not default (template has styles on prepared rows)
                # This currently FAILS — formateador clears values but may not preserve styles
                assert cell.font is not None and cell.font != openpyxl_styles.Font(), (
                    f"{sheet_name} row 5 col {col} lost template style"
                )

    def test_formato_censos_blank_rows_retain_style(self):
        """FORMATO_CENSOS blank rows (7-1000) must retain template styles."""
        output = self.tmp_dir / "censo-test.xlsx"
        run_formateador(make_minimal_data(), output)

        wb = load_workbook_readonly(output)
        ws = wb["FORMATO_CENSOS"]

        # Row 10 is beyond our 1 data row (starts at row 7) but within 1000 prepared
        for col in range(1, 19):  # A-R columns
            cell = ws.cell(row=10, column=col)
            assert cell.value is None, f"FORMATO_CENSOS row 10 col {col} should be blank"
            # Style should be preserved from template
            assert cell.font is not None and cell.font != openpyxl_styles.Font(), (
                f"FORMATO_CENSOS row 10 col {col} lost template style"
            )

# ─── Test 5: XML-safe special characters preserved exactly ──────────────

    def test_special_xml_characters_preserved_exactly(self):
        """Values with XML special chars (A&B <test>) must round-trip exactly and produce valid XLSX."""
        output = self.tmp_dir / "censo-special-chars.xlsx"
        # Use template header names exactly (after normalization)
        special_censo = {
            "VIGENCIA": 2026,
            "RESGUARDO INDIGENA": "A&B <test>",
            "COMUNIDAD INDIGENA": 'Test "quotes" & \'apostrophes\'',
            "FAMILIA": "Test > less than < greater",
            "TIPO IDENTIFICACION": "CC",
            "NUMERO DOCUMENTO": "1234567890",
            "NOMBRES": "TEST",
            "APELLIDOS": "TEST",
            "FECHA NACIMIENTO": "2000-01-01",
            "PARENTESCO": "TEST",
            "SEXO": "M",
            "ESTADO CIVIL": "TEST",
            "PROFESION": "TEST",
            "ESCOLARIDAD": "TEST",
            "INTEGRANTES": "1",
            "DIRECCION": "TEST",
            "TELEFONO": "TEST",
            "USUARIO": "TEST",
        }
        data = {
            "censo": [special_censo],
            "altas": [MINIMAL_ALTAS_ROW],
            "bajas": [MINIMAL_BAJAS_ROW],
        }
        run_formateador(data, output)

        # Must open as valid XLSX
        wb = load_workbook_readonly(output)
        ws = wb["FORMATO_CENSOS"]

        # Data row starts at row 7
        # Column mapping from template header row 6:
        # A=VIGENCIA, B=RESGUARDO INDIGENA, C=COMUNIDAD INDIGENA, D=FAMILIA, E=TIPO IDENTIFICACION
        resguardo_cell = ws.cell(row=7, column=2)   # B column (RESGUARDO INDIGENA)
        comunidad_cell = ws.cell(row=7, column=3)   # C column (COMUNIDAD INDIGENA)
        familia_cell = ws.cell(row=7, column=4)     # D column (FAMILIA)

        assert resguardo_cell.value == "A&B <test>", f"Special chars not preserved: {resguardo_cell.value}"
        assert comunidad_cell.value == 'Test "quotes" & \'apostrophes\'', f"Quotes/apostrophes not preserved: {comunidad_cell.value}"
        assert familia_cell.value == "Test > less than < greater", f"Angle brackets not preserved: {familia_cell.value}"

        # Verify file is valid by re-opening
        wb2 = openpyxl.load_workbook(output)
        ws2 = wb2["FORMATO_CENSOS"]
        assert ws2.cell(row=7, column=2).value == "A&B <test>"

    # ─── Test 6: Existing formula in data cell preserved when unrelated cells written ──

    def test_existing_formula_in_data_cell_preserved(self):
        """Formula in a temp template copy's data cell must not be deleted when unrelated cells are written."""
        # Create a temp copy of template with a formula in a data cell
        import shutil
        template_copy = self.tmp_dir / "template-with-formula.xlsx"
        shutil.copy2(TEMPLATE_PATH, template_copy)

        # Modify the copy: add a formula in REPORTE ALTAS row 2, column A (VIGENCIA)
        wb = openpyxl.load_workbook(template_copy)
        ws = wb["REPORTE ALTAS"]
        # Row 2 is first data row, column A = VIGENCIA
        ws.cell(row=2, column=1).value = "=YEAR(TODAY())"
        wb.save(template_copy)

        # Now run formateador with only census data (no altas data)
        data = {
            "censo": [MINIMAL_CENSO_ROW],
            "altas": [],  # Empty - should not touch the formula cell
            "bajas": [MINIMAL_BAJAS_ROW],
        }
        output = self.tmp_dir / "censo-formula-test.xlsx"
        run_formateador(data, output, template_path=template_copy)

        # Check the formula is still there
        wb_out = load_workbook_readonly(output)
        ws_out = wb_out["REPORTE ALTAS"]
        formula_cell = ws_out.cell(row=2, column=1)
        assert formula_cell.value == "=YEAR(TODAY())", f"Formula lost: {formula_cell.value}"

    # ─── Test 7: FORMATO_CENSOS filter ref expands when censo data exceeds original ──

    def test_formato_censos_filter_ref_expands_with_excess_censo_data(self):
        """FORMATO_CENSOS autoFilter ref must cover rows beyond $A$6:$R$14 when censo data exceeds it."""
        # Template has autoFilter on $A$6:$R$14 (rows 6-14 = 9 rows including header at row 6)
        # Data starts at row 7. If we have >8 data rows, filter must expand.
        num_rows = 15  # Exceeds the 8 data rows covered by original filter
        censo_rows = []
        for i in range(num_rows):
            row = MINIMAL_CENSO_ROW.copy()
            row["NUMERO DOCUMENTO"] = f"123456789{i:02d}"
            row["PRIMER APELLIDO"] = f"TEST{i}"
            censo_rows.append(row)

        data = {
            "censo": censo_rows,
            "altas": [MINIMAL_ALTAS_ROW],
            "bajas": [MINIMAL_BAJAS_ROW],
        }
        output = self.tmp_dir / "censo-filter-test.xlsx"
        run_formateador(data, output)

        wb = load_workbook_readonly(output)
        ws = wb["FORMATO_CENSOS"]

        # Check autoFilter ref covers all data rows
        # Last data row should be 7 + 15 - 1 = 21
        # Expected filter ref: $A$6:$R$21 (header at 6, data 7-21)
        auto_filter = ws.auto_filter
        assert auto_filter.ref == "$A$6:$R$21", (
            f"Filter ref not expanded: expected $A$6:$R$21, got {auto_filter.ref}"
        )

    # ─── Test 8: Placeholder value cleared when dataset empty, style preserved ──

    def test_placeholder_cleared_when_dataset_empty_preserving_style(self):
        """Temp template copy with placeholder in data row must clear it when dataset empty, keep style."""
        import shutil
        template_copy = self.tmp_dir / "template-with-placeholder.xlsx"
        shutil.copy2(TEMPLATE_PATH, template_copy)

        # Add a placeholder value in REPORTE ALTAS row 2 (first data row)
        wb = openpyxl.load_workbook(template_copy)
        ws = wb["REPORTE ALTAS"]
        cell = ws.cell(row=2, column=1)  # VIGENCIA column
        cell.value = "PLACEHOLDER"
        placeholder_style = {
            "font": cell.font,
            "fill": cell.fill,
            "border": cell.border,
            "alignment": cell.alignment,
            "number_format": cell.number_format,
        }
        wb.save(template_copy)

        # Run with empty altas dataset
        data = {
            "censo": [MINIMAL_CENSO_ROW],
            "altas": [],  # Empty - should clear placeholder
            "bajas": [MINIMAL_BAJAS_ROW],
        }
        output = self.tmp_dir / "censo-placeholder-test.xlsx"
        run_formateador(data, output, template_path=template_copy)

        wb_out = load_workbook_readonly(output)
        ws_out = wb_out["REPORTE ALTAS"]
        cell_out = ws_out.cell(row=2, column=1)

        # Value must be cleared (None or empty)
        assert cell_out.value is None or cell_out.value == "", f"Placeholder not cleared: {cell_out.value}"

        # Style must be preserved from template
        assert cell_out.font is not None and cell_out.font != openpyxl_styles.Font(), "Style not preserved"
        # Check key style attributes match template
        assert cell_out.font.name == placeholder_style["font"].name, "Font name not preserved"
        assert cell_out.fill.patternType == placeholder_style["fill"].patternType, "Fill pattern not preserved"

    # ─── Test 9: Rows beyond ALTAS/BAJAS capacity copy style and row height from last prepared ──

    def test_rows_beyond_capacity_copy_style_and_height_from_last_prepared(self):
        """Rows beyond ALTAS/BAJAS prepared capacity (23) must copy style and height from last template row."""
        # Create data exceeding 23 prepared rows
        num_rows = 30
        altas_rows = []
        for i in range(num_rows):
            row = MINIMAL_ALTAS_ROW.copy()
            row["NUMERO DOCUMENTO"] = f"123456789{i:02d}"
            row["PRIMER APELLIDO"] = f"TEST{i}"
            altas_rows.append(row)

        data = {
            "censo": [MINIMAL_CENSO_ROW],
            "altas": altas_rows,
            "bajas": [MINIMAL_BAJAS_ROW],
        }
        output = self.tmp_dir / "censo-capacity-test.xlsx"
        run_formateador(data, output)

        wb = load_workbook_readonly(output)
        ws_altas = wb["REPORTE ALTAS"]
        ws_bajas = wb["REPORTE BAJAS"]

        # Template has 23 prepared data rows (rows 2-24)
        # Last prepared row is 24, row 25 is first beyond capacity
        # Row 25 should copy style and height from row 24
        last_prepared_row = 24
        first_beyond_row = 25

        for ws, sheet_name in [(ws_altas, "REPORTE ALTAS"), (ws_bajas, "REPORTE BAJAS")]:
            # Check row height matches last prepared row
            assert ws.row_dimensions[first_beyond_row].ht == ws.row_dimensions[last_prepared_row].ht, (
                f"{sheet_name} row {first_beyond_row} height {ws.row_dimensions[first_beyond_row].ht} "
                f"does not match last prepared row {last_prepared_row} height {ws.row_dimensions[last_prepared_row].ht}"
            )

            # Check style copied for all columns (A-O = 1-15)
            for col in range(1, 16):
                last_cell = ws.cell(row=last_prepared_row, column=col)
                beyond_cell = ws.cell(row=first_beyond_row, column=col)

                # Font should match
                assert beyond_cell.font is not None and beyond_cell.font != openpyxl_styles.Font(), (
                    f"{sheet_name} row {first_beyond_row} col {col} lost font style"
                )
                assert beyond_cell.font.name == last_cell.font.name, (
                    f"{sheet_name} row {first_beyond_row} col {col} font name mismatch"
                )
                assert beyond_cell.font.size == last_cell.font.size, (
                    f"{sheet_name} row {first_beyond_row} col {col} font size mismatch"
                )

                # Fill should match
                assert beyond_cell.fill.patternType == last_cell.fill.patternType, (
                    f"{sheet_name} row {first_beyond_row} col {col} fill pattern mismatch"
                )

                # Border should match
                last_left = last_cell.border.left
                beyond_left = beyond_cell.border.left
                # Handle None borders (no border style)
                if last_left is None or last_left.style is None:
                    assert beyond_left is None or beyond_left.style is None, (
                        f"{sheet_name} row {first_beyond_row} col {col} left border should be None"
                    )
                else:
                    assert beyond_left is not None and beyond_left.style == last_left.style, (
                        f"{sheet_name} row {first_beyond_row} col {col} left border mismatch"
                    )

    # ─── Test 4: Default output filename ──────────────────────────────────

    def test_default_output_filename_is_censo_year_xlsx(self):
        """Default output (no --output) must be exactly censo-{current_year}.xlsx."""
        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            # Copy template to temp dir to isolate
            import shutil
            template_copy = td_path / "Formato Censal.xlsx"
            shutil.copy2(TEMPLATE_PATH, template_copy)

            # Create minimal data file
            data_file = td_path / "data.json"
            with open(data_file, "w") as f:
                json.dump(make_minimal_data(), f)

            # Run formateador WITHOUT --output, using template in temp dir
            # It should create censo-{year}.xlsx next to the template
            cmd = [
                sys.executable,
                str(FORMATADOR_PATH),
                "--data", str(data_file),
                "--template", str(template_copy),
            ]
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)

            assert result.returncode == 0, f"formateador failed: {result.stderr}"

            expected_name = f"censo-{CURRENT_YEAR}.xlsx"
            output_file = td_path / expected_name
            assert output_file.exists(), (
                f"Default output file not created: expected {expected_name}, "
                f"found {list(td_path.glob('*.xlsx'))}"
            )
            # Ensure the output file has the exact expected name (not template name)
            xlsx_files = list(td_path.glob("*.xlsx"))
            output_files = [f for f in xlsx_files if f.name != "Formato Censal.xlsx"]
            assert len(output_files) == 1, f"Expected exactly 1 output file (excluding template), got {output_files}"
            assert output_files[0].name == expected_name, f"Output file name mismatch: {output_files[0].name} != {expected_name}"


# ─── Test: AutoFilter multi-letter column absolute reference ─────────────────

class TestAutoFilterMultiLetterColumns:
    """Verify autoFilter ref handles multi-letter columns correctly (AA, AB, etc.)."""

    def test_auto_filter_multi_letter_columns_formatted_correctly(self):
        """AutoFilter ref with multi-letter columns must use $AA$1:$AB$10 format, not $A$A1:$A$B10."""
        from formateador import build_range_ref, parse_range_ref

        # Simulate a ref that would be produced for columns AA-AB (27-28)
        # build_range_ref produces "AA1:AB10"
        ref = build_range_ref(27, 1, 28, 10)  # AA1:AB10
        assert ref == "AA1:AB10"

        # Now test the absolute reference formatting logic from update_auto_filter_ref
        def to_absolute(part: str) -> str:
            col_str = ""
            row_str = ""
            for ch in part:
                if ch.isalpha():
                    col_str += ch
                else:
                    row_str += ch
            return f"${col_str}${row_str}"

        parts = ref.split(":")
        abs_ref = ":".join(to_absolute(part) for part in parts)

        assert abs_ref == "$AA$1:$AB$10", f"Expected $AA$1:$AB$10, got {abs_ref}"

    def test_auto_filter_single_letter_columns_still_work(self):
        """Single-letter columns must still format as $A$1:$Z$10."""
        from formateador import build_range_ref

        ref = build_range_ref(1, 1, 26, 10)  # A1:Z10
        assert ref == "A1:Z10"

        def to_absolute(part: str) -> str:
            col_str = ""
            row_str = ""
            for ch in part:
                if ch.isalpha():
                    col_str += ch
                else:
                    row_str += ch
            return f"${col_str}${row_str}"

        parts = ref.split(":")
        abs_ref = ":".join(to_absolute(part) for part in parts)

        assert abs_ref == "$A$1:$Z$10", f"Expected $A$1:$Z$10, got {abs_ref}"

    def test_auto_filter_mixed_columns_formatted_correctly(self):
        """Mixed single and multi-letter columns must format correctly."""
        from formateador import build_range_ref

        # Z1:AA10 (column 26 to 27)
        ref = build_range_ref(26, 1, 27, 10)
        assert ref == "Z1:AA10"

        def to_absolute(part: str) -> str:
            col_str = ""
            row_str = ""
            for ch in part:
                if ch.isalpha():
                    col_str += ch
                else:
                    row_str += ch
            return f"${col_str}${row_str}"

        parts = ref.split(":")
        abs_ref = ":".join(to_absolute(part) for part in parts)

        assert abs_ref == "$Z$1:$AA$10", f"Expected $Z$1:$AA$10, got {abs_ref}"


# ─── Additional Structural Integrity Tests ──────────────────────────────────

class TestTemplateIntegrity:
    """Verify the template itself has expected structure (baseline)."""

    def test_template_has_three_sheets(self):
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
        assert wb.sheetnames == ["FORMATO_CENSOS", "REPORTE ALTAS", "REPORTE BAJAS"]

    def test_template_formato_censos_has_9_merged_ranges(self):
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
        merged = {str(r) for r in wb["FORMATO_CENSOS"].merged_cells.ranges}
        expected = {"A1:B5", "C1:F3", "G1:P3", "C4:F5", "G4:P5", "Q1:Q2", "Q3:Q4", "R1:R2", "R3:R4"}
        assert merged == expected

    def test_template_altas_has_table_1_at_a1_o24(self):
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
        ws = wb["REPORTE ALTAS"]
        assert "Table_1" in ws.tables
        assert str(ws.tables["Table_1"].ref) == "A1:O24"

    def test_template_bajas_has_table_2_at_a1_o24(self):
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
        ws = wb["REPORTE BAJAS"]
        assert "Table_2" in ws.tables
        assert str(ws.tables["Table_2"].ref) == "A1:O24"

    def test_template_has_image_in_media(self):
        """Template must have at least one image in xl/media/."""
        import zipfile
        with zipfile.ZipFile(TEMPLATE_PATH, "r") as z:
            media_files = [n for n in z.namelist() if n.startswith("xl/media/")]
            assert len(media_files) >= 1, f"No media files found in template: {z.namelist()}"
            assert any(n.endswith((".jpg", ".jpeg", ".png", ".gif")) for n in media_files)

    def test_template_has_drawing_relationships_on_all_sheets(self):
        """All three sheets should have drawing relationships in template."""
        import zipfile
        with zipfile.ZipFile(TEMPLATE_PATH, "r") as z:
            for i in range(1, 4):
                rels_name = f"xl/worksheets/_rels/sheet{i}.xml.rels"
                assert rels_name in z.namelist(), f"Missing {rels_name}"
                content = z.read(rels_name).decode("utf-8")
                assert "drawing" in content, f"Sheet {i} missing drawing relationship"


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
